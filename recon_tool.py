"""
Reconciliation and Validation Tool - Main Orchestrator

This is the main entry point for the reconciliation and validation tool.
It reads configuration from Excel, executes rules, and generates timestamped output.
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Import our modules
from recon_engine import ReconciliationEngine, ReconciliationSummary
from validation_engine import ValidationEngine, ValidationSummary
from genai_helper import GenAIHelper, GenAIConfig, profile_dataframe


@dataclass
class ExecutionLog:
    """Store execution log entries."""
    entries: List[Dict] = field(default_factory=list)
    
    def log(self, level: str, component: str, message: str):
        """Add a log entry."""
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "level": level,
            "component": component,
            "message": message
        }
        self.entries.append(entry)
        print(f"[{entry['timestamp']}] [{level}] [{component}] {message}")
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert log entries to DataFrame."""
        return pd.DataFrame(self.entries)


class ReconTool:
    """
    Main orchestrator for the reconciliation and validation tool.
    """
    
    def __init__(self, config_path: str):
        """
        Initialize the tool with a configuration file path.
        
        Args:
            config_path: Path to the Excel configuration file
        """
        self.config_path = config_path
        self.base_dir = os.path.dirname(os.path.abspath(config_path))
        self.logger = ExecutionLog()
        self.config: Dict[str, Any] = {}
        self.column_mappings: List[Dict] = []
        self.recon_rules: List[Dict] = []
        self.validation_rules: List[Dict] = []
        self.recon_summary: Optional[ReconciliationSummary] = None
        self.validation_summary: Optional[ValidationSummary] = None
        self.genai_helper: Optional[GenAIHelper] = None
        self.execution_timestamp = datetime.now()
        
        # Style definitions for Excel output
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def load_configuration(self) -> bool:
        """
        Load configuration from the Excel file.
        
        Returns:
            True if configuration loaded successfully
        """
        self.logger.log("INFO", "Config", f"Loading configuration from {self.config_path}")
        
        try:
            # Load Config sheet
            config_df = pd.read_excel(self.config_path, sheet_name="Config", header=2)
            config_df.columns = ["section", "parameter", "value"]
            config_df = config_df.dropna(subset=["parameter"])
            
            for _, row in config_df.iterrows():
                param = str(row["parameter"]).strip()
                value = row["value"]
                self.config[param] = value
            
            self.logger.log("INFO", "Config", f"Loaded {len(self.config)} configuration parameters")
            
            # Load Column Mappings
            try:
                mappings_df = pd.read_excel(self.config_path, sheet_name="ColumnMappings", header=2)
                mappings_df.columns = ["mapping_id", "source1_column", "source2_column", "description"]
                self.column_mappings = mappings_df.dropna(subset=["source1_column"]).to_dict("records")
                self.logger.log("INFO", "Config", f"Loaded {len(self.column_mappings)} column mappings")
            except Exception as e:
                self.logger.log("WARNING", "Config", f"Could not load column mappings: {e}")
            
            # Load Reconciliation Rules
            try:
                recon_df = pd.read_excel(self.config_path, sheet_name="ReconciliationRules")
                recon_df.columns = [c.lower().replace(" ", "_").replace("(", "").replace(")", "") 
                                   for c in recon_df.columns]
                self.recon_rules = recon_df.dropna(subset=["rule_id"]).to_dict("records")
                self.logger.log("INFO", "Config", f"Loaded {len(self.recon_rules)} reconciliation rules")
            except Exception as e:
                self.logger.log("WARNING", "Config", f"Could not load reconciliation rules: {e}")
            
            # Load Validation Rules
            try:
                valid_df = pd.read_excel(self.config_path, sheet_name="ValidationRules")
                valid_df.columns = [c.lower().replace(" ", "_").replace("(", "").replace(")", "") 
                                   for c in valid_df.columns]
                self.validation_rules = valid_df.dropna(subset=["rule_id"]).to_dict("records")
                self.logger.log("INFO", "Config", f"Loaded {len(self.validation_rules)} validation rules")
            except Exception as e:
                self.logger.log("WARNING", "Config", f"Could not load validation rules: {e}")
            
            # Initialize GenAI if enabled
            genai_enabled = str(self.config.get("Enable GenAI", "FALSE")).upper() == "TRUE"
            if genai_enabled:
                genai_config = GenAIConfig(
                    enabled=True,
                    provider=self.config.get("API Provider", "Claude"),
                    api_key_env_var=self.config.get("API Key Environment Variable", "ANTHROPIC_API_KEY"),
                    model=self.config.get("Model", "claude-3-5-sonnet-20241022")
                )
                self.genai_helper = GenAIHelper(genai_config, self.logger)
                self.logger.log("INFO", "GenAI", "GenAI integration initialized")
            
            return True
            
        except Exception as e:
            self.logger.log("ERROR", "Config", f"Failed to load configuration: {e}")
            return False
    
    def _resolve_path(self, path: str) -> str:
        """Resolve a relative path to absolute path."""
        if path.startswith("./"):
            return os.path.join(self.base_dir, path[2:])
        elif path.startswith("../"):
            return os.path.join(self.base_dir, path)
        elif not os.path.isabs(path):
            return os.path.join(self.base_dir, path)
        return path
    
    def run_reconciliation(self) -> ReconciliationSummary:
        """
        Execute all reconciliation rules.
        
        Returns:
            ReconciliationSummary with results
        """
        self.logger.log("INFO", "Reconciliation", "Starting reconciliation checks")
        
        engine = ReconciliationEngine(self.config, self.logger)
        
        # Load data sources
        source1_path = self._resolve_path(self.config.get("Source 1 Path", ""))
        source2_path = self._resolve_path(self.config.get("Source 2 Path", ""))
        
        if source1_path and os.path.exists(source1_path):
            engine.load_data_source(
                "source1",
                source1_path,
                self.config.get("Source 1 Type", "CSV"),
                self.config.get("Source 1 Encoding", "utf-8"),
                self.config.get("Source 1 Delimiter", ",")
            )
        
        if source2_path and os.path.exists(source2_path):
            engine.load_data_source(
                "source2",
                source2_path,
                self.config.get("Source 2 Type", "CSV"),
                self.config.get("Source 2 Encoding", "utf-8"),
                self.config.get("Source 2 Delimiter", ",")
            )
        
        # Set column mappings
        engine.set_column_mappings(self.column_mappings)
        
        # Process natural language rules if GenAI is available
        if self.genai_helper and self.genai_helper.is_available():
            self._process_nl_rules("reconciliation")
        
        # Run rules
        self.recon_summary = engine.run_all_rules(self.recon_rules)
        
        self.logger.log("INFO", "Reconciliation", 
                       f"Completed: {self.recon_summary.rules_passed} passed, "
                       f"{self.recon_summary.rules_failed} failed")
        
        return self.recon_summary
    
    def run_validation(self) -> ValidationSummary:
        """
        Execute all validation rules.
        
        Returns:
            ValidationSummary with results
        """
        self.logger.log("INFO", "Validation", "Starting validation checks")
        
        engine = ValidationEngine(self.config, self.logger)
        
        # Load validation data source
        valid_path = self._resolve_path(self.config.get("Validation Data Path", ""))
        
        if valid_path and os.path.exists(valid_path):
            engine.load_data_source(
                "validation_data",
                valid_path,
                "CSV",
                "utf-8",
                ","
            )
        
        # Also load source1 for validation if needed
        source1_path = self._resolve_path(self.config.get("Source 1 Path", ""))
        if source1_path and os.path.exists(source1_path):
            engine.load_data_source(
                "source1",
                source1_path,
                self.config.get("Source 1 Type", "CSV"),
                self.config.get("Source 1 Encoding", "utf-8"),
                self.config.get("Source 1 Delimiter", ",")
            )
        
        # Process natural language rules if GenAI is available
        if self.genai_helper and self.genai_helper.is_available():
            self._process_nl_rules("validation")
        
        # Run rules
        self.validation_summary = engine.run_all_rules(self.validation_rules)
        
        self.logger.log("INFO", "Validation", 
                       f"Completed: {self.validation_summary.rules_passed} passed, "
                       f"{self.validation_summary.rules_failed} failed")
        
        return self.validation_summary
    
    def _process_nl_rules(self, rule_type: str):
        """Process natural language rules using GenAI."""
        rules = self.recon_rules if rule_type == "reconciliation" else self.validation_rules
        nl_column = "natural_language_rule"
        
        for rule in rules:
            nl_rule = rule.get(nl_column, "")
            if nl_rule and str(nl_rule).strip() and str(nl_rule).lower() != "nan":
                self.logger.log("INFO", "GenAI", f"Processing NL rule: {nl_rule[:50]}...")
                parsed = self.genai_helper.parse_natural_language_rule(nl_rule, rule_type)
                if parsed:
                    # Update rule with parsed values (don't override existing values)
                    for key, value in parsed.items():
                        if key not in rule or not rule[key] or str(rule[key]).lower() == "nan":
                            rule[key] = value
                    self.logger.log("INFO", "GenAI", f"Successfully parsed NL rule")
    
    def generate_output(self) -> str:
        """
        Generate timestamped Excel output file with all results.
        
        Returns:
            Path to the generated output file
        """
        timestamp_str = self.execution_timestamp.strftime("%Y%m%d_%H%M%S")
        project_name = self.config.get("Project Name", "reconciliation").replace(" ", "_")
        output_dir = self._resolve_path(self.config.get("Output Directory", "./output"))
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        output_filename = f"{project_name}_{timestamp_str}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        self.logger.log("INFO", "Output", f"Generating output file: {output_path}")
        
        # Copy the config file as base
        shutil.copy(self.config_path, output_path)
        
        # Load the workbook for modification
        wb = load_workbook(output_path)
        
        # Update Output Summary sheet
        self._write_summary_sheet(wb)
        
        # Write Reconciliation Results
        self._write_recon_results(wb)
        
        # Write Validation Results
        self._write_validation_results(wb)
        
        # Write Execution Log
        self._write_execution_log(wb)
        
        # Save the workbook
        wb.save(output_path)
        
        self.logger.log("INFO", "Output", f"Output file generated successfully: {output_path}")
        
        return output_path
    
    def _write_summary_sheet(self, wb):
        """Write the output summary sheet."""
        ws = wb["OutputSummary"]
        
        # Calculate metrics
        recon_match_rate = 0
        if self.recon_summary and self.recon_summary.total_records_source1 > 0:
            recon_match_rate = (self.recon_summary.matched_records / 
                              self.recon_summary.total_records_source1 * 100)
        
        valid_pass_rate = 0
        if self.validation_summary and self.validation_summary.total_records > 0:
            valid_pass_rate = (self.validation_summary.records_passed / 
                             self.validation_summary.total_records * 100)
        
        # Define summary data
        summary_data = [
            ("Execution Timestamp", self.execution_timestamp.strftime("%Y-%m-%d %H:%M:%S"), "COMPLETE", ""),
            ("Total Rules Executed", 
             (self.recon_summary.rules_executed if self.recon_summary else 0) + 
             (self.validation_summary.rules_executed if self.validation_summary else 0),
             "", ""),
            ("Reconciliation Rules", self.recon_summary.rules_executed if self.recon_summary else 0, "", ""),
            ("Validation Rules", self.validation_summary.rules_executed if self.validation_summary else 0, "", ""),
            ("", "", "", ""),
            ("RECONCILIATION SUMMARY", "", "", ""),
            ("Total Records (Source 1)", self.recon_summary.total_records_source1 if self.recon_summary else 0, "", ""),
            ("Total Records (Source 2)", self.recon_summary.total_records_source2 if self.recon_summary else 0, "", ""),
            ("Matched Records", self.recon_summary.matched_records if self.recon_summary else 0, 
             "PASS" if self.recon_summary and self.recon_summary.unmatched_source1 == 0 else "FAIL", ""),
            ("Unmatched in Source 1", self.recon_summary.unmatched_source1 if self.recon_summary else 0,
             "PASS" if self.recon_summary and self.recon_summary.unmatched_source1 == 0 else "FAIL", ""),
            ("Unmatched in Source 2", self.recon_summary.unmatched_source2 if self.recon_summary else 0,
             "PASS" if self.recon_summary and self.recon_summary.unmatched_source2 == 0 else "FAIL", ""),
            ("Value Discrepancies", self.recon_summary.value_discrepancies if self.recon_summary else 0,
             "PASS" if self.recon_summary and self.recon_summary.value_discrepancies == 0 else "FAIL", ""),
            ("Match Rate (%)", f"{recon_match_rate:.2f}%", 
             "PASS" if recon_match_rate >= 95 else "WARNING" if recon_match_rate >= 90 else "FAIL", ""),
            ("", "", "", ""),
            ("VALIDATION SUMMARY", "", "", ""),
            ("Total Records Validated", self.validation_summary.total_records if self.validation_summary else 0, "", ""),
            ("Records Passed", self.validation_summary.records_passed if self.validation_summary else 0, "", ""),
            ("Records with Errors", self.validation_summary.records_with_errors if self.validation_summary else 0,
             "PASS" if self.validation_summary and self.validation_summary.records_with_errors == 0 else "FAIL", ""),
            ("Records with Warnings", self.validation_summary.records_with_warnings if self.validation_summary else 0,
             "PASS" if self.validation_summary and self.validation_summary.records_with_warnings == 0 else "WARNING", ""),
            ("Pass Rate (%)", f"{valid_pass_rate:.2f}%",
             "PASS" if valid_pass_rate >= 95 else "WARNING" if valid_pass_rate >= 90 else "FAIL", ""),
        ]
        
        for row_idx, (metric, value, status, details) in enumerate(summary_data, 2):
            ws.cell(row=row_idx, column=1, value=metric).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=value).border = self.thin_border
            cell_status = ws.cell(row=row_idx, column=3, value=status)
            cell_status.border = self.thin_border
            ws.cell(row=row_idx, column=4, value=details).border = self.thin_border
            
            # Apply conditional formatting
            if status == "PASS":
                cell_status.fill = self.pass_fill
            elif status == "FAIL":
                cell_status.fill = self.fail_fill
            elif status == "WARNING":
                cell_status.fill = self.warning_fill
    
    def _write_recon_results(self, wb):
        """Write reconciliation results to the sheet."""
        ws = wb["ReconResults"]
        
        if not self.recon_summary or not self.recon_summary.results:
            ws.cell(row=2, column=1, value="No reconciliation results")
            return
        
        for row_idx, result in enumerate(self.recon_summary.results, 2):
            ws.cell(row=row_idx, column=1, value=result.rule_id).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=result.rule_name).border = self.thin_border
            ws.cell(row=row_idx, column=3, value=result.record_key).border = self.thin_border
            ws.cell(row=row_idx, column=4, value=str(result.source1_value)).border = self.thin_border
            ws.cell(row=row_idx, column=5, value=str(result.source2_value)).border = self.thin_border
            ws.cell(row=row_idx, column=6, value=str(result.difference)).border = self.thin_border
            
            status_cell = ws.cell(row=row_idx, column=7, value=result.status)
            status_cell.border = self.thin_border
            if result.status == "PASS":
                status_cell.fill = self.pass_fill
            elif result.status == "FAIL":
                status_cell.fill = self.fail_fill
            elif result.status == "WARNING":
                status_cell.fill = self.warning_fill
            
            ws.cell(row=row_idx, column=8, value=result.severity).border = self.thin_border
            ws.cell(row=row_idx, column=9, value=result.details).border = self.thin_border
    
    def _write_validation_results(self, wb):
        """Write validation results to the sheet."""
        ws = wb["ValidationResults"]
        
        if not self.validation_summary or not self.validation_summary.results:
            ws.cell(row=2, column=1, value="No validation results")
            return
        
        for row_idx, result in enumerate(self.validation_summary.results, 2):
            ws.cell(row=row_idx, column=1, value=result.rule_id).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=result.rule_name).border = self.thin_border
            ws.cell(row=row_idx, column=3, value=result.record_key).border = self.thin_border
            ws.cell(row=row_idx, column=4, value=result.column).border = self.thin_border
            ws.cell(row=row_idx, column=5, value=str(result.value)).border = self.thin_border
            ws.cell(row=row_idx, column=6, value=str(result.expected)).border = self.thin_border
            
            status_cell = ws.cell(row=row_idx, column=7, value=result.status)
            status_cell.border = self.thin_border
            if result.status == "PASS":
                status_cell.fill = self.pass_fill
            elif result.status == "FAIL":
                status_cell.fill = self.fail_fill
            elif result.status == "WARNING":
                status_cell.fill = self.warning_fill
            
            ws.cell(row=row_idx, column=8, value=result.severity).border = self.thin_border
            ws.cell(row=row_idx, column=9, value=result.details).border = self.thin_border
    
    def _write_execution_log(self, wb):
        """Write execution log to the sheet."""
        ws = wb["ExecutionLog"]
        
        for row_idx, entry in enumerate(self.logger.entries, 2):
            ws.cell(row=row_idx, column=1, value=entry["timestamp"]).border = self.thin_border
            
            level_cell = ws.cell(row=row_idx, column=2, value=entry["level"])
            level_cell.border = self.thin_border
            if entry["level"] == "ERROR":
                level_cell.fill = self.fail_fill
            elif entry["level"] == "WARNING":
                level_cell.fill = self.warning_fill
            
            ws.cell(row=row_idx, column=3, value=entry["component"]).border = self.thin_border
            ws.cell(row=row_idx, column=4, value=entry["message"]).border = self.thin_border
    
    def run(self) -> str:
        """
        Execute the full reconciliation and validation workflow.
        
        Returns:
            Path to the generated output file
        """
        self.logger.log("INFO", "Main", "=" * 60)
        self.logger.log("INFO", "Main", "RECONCILIATION AND VALIDATION TOOL")
        self.logger.log("INFO", "Main", "=" * 60)
        
        # Load configuration
        if not self.load_configuration():
            self.logger.log("ERROR", "Main", "Failed to load configuration. Aborting.")
            return ""
        
        # Run reconciliation
        if self.recon_rules:
            self.run_reconciliation()
        else:
            self.logger.log("INFO", "Main", "No reconciliation rules defined. Skipping.")
        
        # Run validation
        if self.validation_rules:
            self.run_validation()
        else:
            self.logger.log("INFO", "Main", "No validation rules defined. Skipping.")
        
        # Generate output
        output_path = self.generate_output()
        
        self.logger.log("INFO", "Main", "=" * 60)
        self.logger.log("INFO", "Main", "EXECUTION COMPLETE")
        self.logger.log("INFO", "Main", f"Output file: {output_path}")
        self.logger.log("INFO", "Main", "=" * 60)
        
        return output_path


def main():
    """Main entry point for the tool."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Reconciliation and Validation Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python recon_tool.py config.xlsx
  python recon_tool.py /path/to/config.xlsx
        """
    )
    parser.add_argument(
        "config",
        help="Path to the Excel configuration file"
    )
    parser.add_argument(
        "--enable-genai",
        action="store_true",
        help="Enable GenAI features (requires API key)"
    )
    
    args = parser.parse_args()
    
    # Check if config file exists
    if not os.path.exists(args.config):
        print(f"Error: Configuration file not found: {args.config}")
        sys.exit(1)
    
    # Run the tool
    tool = ReconTool(args.config)
    output_path = tool.run()
    
    if output_path:
        print(f"\nSuccess! Output file: {output_path}")
        sys.exit(0)
    else:
        print("\nExecution failed. Check the logs for details.")
        sys.exit(1)


if __name__ == "__main__":
    main()
