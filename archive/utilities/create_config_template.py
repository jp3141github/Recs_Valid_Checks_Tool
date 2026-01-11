"""
Excel Configuration Template Generator

This script creates the Excel configuration file that serves as both
the input configuration and the output container for the reconciliation tool.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

def create_config_template():
    """Create the Excel configuration template with all required sheets."""
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========================================================================
    # SHEET 1: Config (Global Settings)
    # =========================================================================
    ws_config = wb.active
    ws_config.title = "Config"
    
    config_data = [
        ["RECONCILIATION TOOL - CONFIGURATION", "", ""],
        ["", "", ""],
        ["SECTION", "PARAMETER", "VALUE"],
        ["General", "Project Name", "Financial Reconciliation Q4 2025"],
        ["General", "Run Description", "Monthly transaction reconciliation"],
        ["General", "Output Directory", "./output"],
        ["General", "Log Level", "INFO"],
        ["", "", ""],
        ["Data Sources", "Source 1 Path", "./data/source1_transactions.csv"],
        ["Data Sources", "Source 1 Type", "CSV"],
        ["Data Sources", "Source 1 Encoding", "utf-8"],
        ["Data Sources", "Source 1 Delimiter", ","],
        ["Data Sources", "Source 2 Path", "./data/source2_bank_records.csv"],
        ["Data Sources", "Source 2 Type", "CSV"],
        ["Data Sources", "Source 2 Encoding", "utf-8"],
        ["Data Sources", "Source 2 Delimiter", ","],
        ["", "", ""],
        ["Reference Data", "Categories Path", "./data/ref_categories.csv"],
        ["Reference Data", "Regions Path", "./data/ref_regions.csv"],
        ["Reference Data", "Currencies Path", "./data/ref_currencies.csv"],
        ["Reference Data", "Accounts Path", "./data/ref_accounts.csv"],
        ["", "", ""],
        ["Validation Data", "Validation Data Path", "./data/validation_test_data.csv"],
        ["", "", ""],
        ["GenAI Settings", "Enable GenAI", "FALSE"],
        ["GenAI Settings", "API Provider", "Claude"],
        ["GenAI Settings", "API Key Environment Variable", "ANTHROPIC_API_KEY"],
        ["GenAI Settings", "Model", "claude-3-5-sonnet-20241022"],
        ["", "", ""],
        ["Thresholds", "Amount Tolerance (%)", "0.01"],
        ["Thresholds", "Date Tolerance (days)", "0"],
        ["Thresholds", "Max Errors Before Abort", "1000"],
    ]
    
    for row_idx, row_data in enumerate(config_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
            elif value in ["General", "Data Sources", "Reference Data", "Validation Data", "GenAI Settings", "Thresholds"]:
                cell.fill = section_fill
                cell.font = Font(bold=True)
    
    ws_config.column_dimensions['A'].width = 20
    ws_config.column_dimensions['B'].width = 30
    ws_config.column_dimensions['C'].width = 50
    
    # =========================================================================
    # SHEET 2: Column Mappings
    # =========================================================================
    ws_mappings = wb.create_sheet("ColumnMappings")
    
    mapping_data = [
        ["COLUMN MAPPINGS - Define how columns relate between sources", "", "", ""],
        ["", "", "", ""],
        ["Mapping ID", "Source 1 Column", "Source 2 Column", "Description"],
        ["MAP001", "transaction_id", "ref_id", "Primary key for matching records"],
        ["MAP002", "transaction_date", "posting_date", "Transaction date field"],
        ["MAP003", "amount", "transaction_amount", "Transaction amount"],
        ["MAP004", "category", "trans_type", "Transaction category/type"],
        ["MAP005", "status", "trans_status", "Transaction status"],
        ["MAP006", "customer_name", "client_name", "Customer/Client name"],
        ["MAP007", "account_number", "client_account", "Account number"],
        ["MAP008", "region", "branch_region", "Region/Branch"],
        ["MAP009", "currency", "currency_code", "Currency code"],
        ["MAP010", "quantity", "units", "Quantity/Units"],
        ["MAP011", "unit_price", "price_per_unit", "Unit price"],
    ]
    
    for row_idx, row_data in enumerate(mapping_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_mappings.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_mappings.column_dimensions['A'].width = 15
    ws_mappings.column_dimensions['B'].width = 25
    ws_mappings.column_dimensions['C'].width = 25
    ws_mappings.column_dimensions['D'].width = 40
    
    # =========================================================================
    # SHEET 3: Reconciliation Rules
    # =========================================================================
    ws_recon = wb.create_sheet("ReconciliationRules")
    
    recon_headers = [
        "Rule ID", "Rule Name", "Active", "Source 1", "Source 2", 
        "Key Column (S1)", "Key Column (S2)", "Check Type", 
        "Compare Column (S1)", "Compare Column (S2)", 
        "Tolerance", "Tolerance Type", "Description", "Natural Language Rule"
    ]
    
    recon_rules = [
        ["RECON001", "Match Transaction IDs", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "transaction_id", "ref_id", "key_match", "", "", "", "", 
         "Verify all transaction IDs exist in both sources", ""],
        ["RECON002", "Amount Reconciliation", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "transaction_id", "ref_id", "value_equals", "amount", "transaction_amount",
         "0.01", "percentage", "Compare transaction amounts with 0.01% tolerance", ""],
        ["RECON003", "Date Reconciliation", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "transaction_id", "ref_id", "value_equals", "transaction_date", "posting_date",
         "0", "days", "Verify transaction dates match exactly", ""],
        ["RECON004", "Status Reconciliation", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "transaction_id", "ref_id", "value_equals", "status", "trans_status",
         "", "", "Verify transaction statuses match", ""],
        ["RECON005", "Customer Name Match", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "transaction_id", "ref_id", "fuzzy_match", "customer_name", "client_name",
         "90", "similarity_score", "Fuzzy match customer names (90% similarity)", ""],
        ["RECON006", "Quantity Check", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "transaction_id", "ref_id", "value_equals", "quantity", "units",
         "", "", "Verify quantities match", ""],
        ["RECON007", "Sum Reconciliation", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "", "", "aggregate_sum", "amount", "transaction_amount",
         "0.01", "percentage", "Compare total sums between sources", ""],
        ["RECON008", "Count Reconciliation", "TRUE", "source1_transactions.csv", "source2_bank_records.csv",
         "", "", "aggregate_count", "", "",
         "", "", "Compare record counts between sources", ""],
    ]
    
    # Write headers
    for col_idx, header in enumerate(recon_headers, 1):
        cell = ws_recon.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Write rules
    for row_idx, rule in enumerate(recon_rules, 2):
        for col_idx, value in enumerate(rule, 1):
            cell = ws_recon.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Set column widths
    col_widths = [12, 25, 8, 30, 30, 18, 18, 15, 20, 20, 10, 15, 45, 50]
    for idx, width in enumerate(col_widths, 1):
        ws_recon.column_dimensions[chr(64 + idx) if idx <= 26 else 'N'].width = width
    
    # =========================================================================
    # SHEET 4: Validation Rules
    # =========================================================================
    ws_valid = wb.create_sheet("ValidationRules")
    
    valid_headers = [
        "Rule ID", "Rule Name", "Active", "Data Source", "Column", 
        "Check Type", "Parameter 1", "Parameter 2", "Severity",
        "Description", "Natural Language Rule"
    ]
    
    valid_rules = [
        ["VAL001", "Amount Positive", "TRUE", "validation_test_data.csv", "amount",
         "greater_than", "0", "", "ERROR", "Amount must be positive", ""],
        ["VAL002", "Percentage Range", "TRUE", "validation_test_data.csv", "percentage",
         "between", "0", "100", "ERROR", "Percentage must be between 0 and 100", ""],
        ["VAL003", "Email Format", "TRUE", "validation_test_data.csv", "email",
         "regex_match", r"^[\w\.-]+@[\w\.-]+\.\w+$", "", "ERROR", "Email must be valid format", ""],
        ["VAL004", "Phone Format", "TRUE", "validation_test_data.csv", "phone",
         "regex_match", r"^\+\d{1,3}-\d{3}-\d{3}-\d{4}$", "", "WARNING", "Phone should match format", ""],
        ["VAL005", "Date Valid", "TRUE", "validation_test_data.csv", "date_field",
         "is_date", "YYYY-MM-DD", "", "ERROR", "Date must be valid format", ""],
        ["VAL006", "Category Valid", "TRUE", "validation_test_data.csv", "category",
         "is_in_list", "Sales,Refund,Transfer,Fee,Interest,Adjustment", "", "ERROR", 
         "Category must be from valid list", ""],
        ["VAL007", "Not Null Amount", "TRUE", "validation_test_data.csv", "amount",
         "not_null", "", "", "ERROR", "Amount cannot be null", ""],
        ["VAL008", "Not Empty Email", "TRUE", "validation_test_data.csv", "email",
         "not_empty", "", "", "ERROR", "Email cannot be empty", ""],
        ["VAL009", "Status Valid", "TRUE", "validation_test_data.csv", "status",
         "is_in_list", "Active,Inactive,Pending,Closed", "", "ERROR", 
         "Status must be from valid list", ""],
        ["VAL010", "Score Range", "TRUE", "validation_test_data.csv", "score",
         "between", "0", "100", "WARNING", "Score should be between 0 and 100", ""],
        ["VAL011", "Unique Record ID", "TRUE", "validation_test_data.csv", "record_id",
         "unique", "", "", "ERROR", "Record ID must be unique", ""],
        ["VAL012", "Amount Data Type", "TRUE", "validation_test_data.csv", "amount",
         "is_numeric", "", "", "ERROR", "Amount must be numeric", ""],
    ]
    
    # Write headers
    for col_idx, header in enumerate(valid_headers, 1):
        cell = ws_valid.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Write rules
    for row_idx, rule in enumerate(valid_rules, 2):
        for col_idx, value in enumerate(rule, 1):
            cell = ws_valid.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Set column widths
    col_widths_v = [12, 20, 8, 30, 15, 15, 45, 15, 10, 40, 50]
    for idx, width in enumerate(col_widths_v, 1):
        ws_valid.column_dimensions[chr(64 + idx) if idx <= 26 else 'K'].width = width
    
    # =========================================================================
    # SHEET 5: Output Summary (Template - will be populated by Python)
    # =========================================================================
    ws_summary = wb.create_sheet("OutputSummary")
    
    summary_headers = ["Metric", "Value", "Status", "Details"]
    summary_template = [
        ["Execution Timestamp", "", "", ""],
        ["Total Rules Executed", "", "", ""],
        ["Reconciliation Rules", "", "", ""],
        ["Validation Rules", "", "", ""],
        ["", "", "", ""],
        ["RECONCILIATION SUMMARY", "", "", ""],
        ["Total Records (Source 1)", "", "", ""],
        ["Total Records (Source 2)", "", "", ""],
        ["Matched Records", "", "", ""],
        ["Unmatched in Source 1", "", "", ""],
        ["Unmatched in Source 2", "", "", ""],
        ["Value Discrepancies", "", "", ""],
        ["Match Rate (%)", "", "", ""],
        ["", "", "", ""],
        ["VALIDATION SUMMARY", "", "", ""],
        ["Total Records Validated", "", "", ""],
        ["Records Passed", "", "", ""],
        ["Records with Errors", "", "", ""],
        ["Records with Warnings", "", "", ""],
        ["Pass Rate (%)", "", "", ""],
    ]
    
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(summary_template, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if value in ["RECONCILIATION SUMMARY", "VALIDATION SUMMARY"]:
                cell.font = Font(bold=True)
                cell.fill = section_fill
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 50
    
    # =========================================================================
    # SHEET 6: Reconciliation Results (Template)
    # =========================================================================
    ws_recon_results = wb.create_sheet("ReconResults")
    
    recon_result_headers = [
        "Rule ID", "Rule Name", "Record Key", "Source 1 Value", "Source 2 Value",
        "Difference", "Status", "Severity", "Details"
    ]
    
    for col_idx, header in enumerate(recon_result_headers, 1):
        cell = ws_recon_results.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    col_widths_r = [12, 25, 25, 20, 20, 15, 12, 10, 50]
    for idx, width in enumerate(col_widths_r, 1):
        ws_recon_results.column_dimensions[chr(64 + idx)].width = width
    
    # =========================================================================
    # SHEET 7: Validation Results (Template)
    # =========================================================================
    ws_valid_results = wb.create_sheet("ValidationResults")
    
    valid_result_headers = [
        "Rule ID", "Rule Name", "Record Key", "Column", "Value",
        "Expected", "Status", "Severity", "Details"
    ]
    
    for col_idx, header in enumerate(valid_result_headers, 1):
        cell = ws_valid_results.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    col_widths_vr = [12, 25, 20, 15, 25, 25, 12, 10, 50]
    for idx, width in enumerate(col_widths_vr, 1):
        ws_valid_results.column_dimensions[chr(64 + idx)].width = width
    
    # =========================================================================
    # SHEET 8: Execution Log (Template)
    # =========================================================================
    ws_log = wb.create_sheet("ExecutionLog")
    
    log_headers = ["Timestamp", "Level", "Component", "Message"]
    
    for col_idx, header in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    ws_log.column_dimensions['A'].width = 25
    ws_log.column_dimensions['B'].width = 10
    ws_log.column_dimensions['C'].width = 20
    ws_log.column_dimensions['D'].width = 80
    
    # =========================================================================
    # SHEET 9: Rule Reference Guide
    # =========================================================================
    ws_guide = wb.create_sheet("RuleGuide")
    
    guide_content = [
        ["RULE REFERENCE GUIDE", "", ""],
        ["", "", ""],
        ["RECONCILIATION CHECK TYPES", "", ""],
        ["Check Type", "Description", "Parameters"],
        ["key_match", "Verify records exist in both sources based on key columns", "Key columns in both sources"],
        ["value_equals", "Compare values between matched records", "Columns to compare, optional tolerance"],
        ["fuzzy_match", "Compare text values with fuzzy matching", "Columns to compare, similarity threshold"],
        ["aggregate_sum", "Compare sum of values between sources", "Column to sum"],
        ["aggregate_count", "Compare record counts between sources", "None"],
        ["aggregate_avg", "Compare average values between sources", "Column to average"],
        ["", "", ""],
        ["VALIDATION CHECK TYPES", "", ""],
        ["Check Type", "Description", "Parameters"],
        ["not_null", "Value must not be null/None", "None"],
        ["not_empty", "Value must not be empty string", "None"],
        ["greater_than", "Value must be greater than parameter", "Threshold value"],
        ["less_than", "Value must be less than parameter", "Threshold value"],
        ["between", "Value must be between two parameters", "Min value, Max value"],
        ["equals", "Value must equal parameter", "Expected value"],
        ["is_in_list", "Value must be in comma-separated list", "Comma-separated valid values"],
        ["regex_match", "Value must match regex pattern", "Regex pattern"],
        ["is_date", "Value must be valid date", "Date format (e.g., YYYY-MM-DD)"],
        ["is_numeric", "Value must be numeric", "None"],
        ["unique", "Values in column must be unique", "None"],
        ["", "", ""],
        ["TOLERANCE TYPES", "", ""],
        ["Type", "Description", "Example"],
        ["percentage", "Tolerance as percentage of value", "0.01 = 0.01% tolerance"],
        ["absolute", "Absolute numeric tolerance", "0.01 = ±0.01 tolerance"],
        ["days", "Date tolerance in days", "1 = ±1 day tolerance"],
        ["similarity_score", "Text similarity threshold (0-100)", "90 = 90% similar"],
        ["", "", ""],
        ["SEVERITY LEVELS", "", ""],
        ["Level", "Description", "Action"],
        ["ERROR", "Critical issue that must be resolved", "Fails the check"],
        ["WARNING", "Non-critical issue for review", "Flags for attention"],
        ["INFO", "Informational finding", "Logged only"],
    ]
    
    for row_idx, row_data in enumerate(guide_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_guide.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.fill = header_fill
                cell.font = header_font
            elif value in ["RECONCILIATION CHECK TYPES", "VALIDATION CHECK TYPES", 
                          "TOLERANCE TYPES", "SEVERITY LEVELS"]:
                cell.font = Font(bold=True)
                cell.fill = section_fill
            elif row_idx in [4, 13, 26, 33]:
                cell.font = header_font
                cell.fill = header_fill
    
    ws_guide.column_dimensions['A'].width = 20
    ws_guide.column_dimensions['B'].width = 50
    ws_guide.column_dimensions['C'].width = 40
    
    # Save the workbook
    output_path = "/home/ubuntu/recon_tool/recon_config_template.xlsx"
    wb.save(output_path)
    print(f"Configuration template created: {output_path}")
    
    return output_path

if __name__ == "__main__":
    create_config_template()
